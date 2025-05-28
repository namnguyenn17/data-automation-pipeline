"""
Report Generation module for the Data Automation Pipeline.

This module handles:
- HTML report generation with interactive charts
- PDF report creation with professional styling
- Excel report with multiple sheets and formatting
- Chart generation and visualization
- Template rendering with Jinja2
"""

import logging
import os
import base64
from io import BytesIO
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
from jinja2 import Environment, FileSystemLoader, Template
# WeasyPrint is disabled due to macOS system library dependencies
WEASYPRINT_AVAILABLE = False
weasyprint = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from scripts.config import Config


class ReportGenerator:
    """Handles report generation in multiple formats."""
    
    def __init__(self):
        """Initialize the report generator."""
        self.logger = logging.getLogger(__name__)
        
        # Setup Jinja2 environment
        self.jinja_env = Environment(
            loader=FileSystemLoader(Config.TEMPLATES_DIR),
            autoescape=True
        )
        
        # Setup plotly configuration
        pio.templates.default = Config.CHART_STYLE
        
    def generate_comprehensive_report(self, data_dict: Dict[str, pd.DataFrame], 
                                    comparison_df: pd.DataFrame,
                                    stats_dict: Dict[str, Dict[str, float]],
                                    timestamp: datetime = None) -> Dict[str, str]:
        """
        Generate comprehensive reports in all formats.
        
        Args:
            data_dict: Dictionary mapping symbols to DataFrames
            comparison_df: Comparison DataFrame
            stats_dict: Dictionary of summary statistics by symbol
            timestamp: Report generation timestamp
            
        Returns:
            Dictionary mapping format to file path
        """
        if timestamp is None:
            timestamp = datetime.now()
        
        report_files = {}
        
        # Generate HTML report
        if 'html' in Config.REPORT_FORMATS:
            html_file = self.generate_html_report(data_dict, comparison_df, stats_dict, timestamp)
            if html_file:
                report_files['html'] = html_file
        
        # Generate PDF report
        if 'pdf' in Config.REPORT_FORMATS:
            pdf_file = self.generate_pdf_report(data_dict, comparison_df, stats_dict, timestamp)
            if pdf_file:
                report_files['pdf'] = pdf_file
        
        # Generate Excel report
        if 'excel' in Config.REPORT_FORMATS:
            excel_file = self.generate_excel_report(data_dict, comparison_df, stats_dict, timestamp)
            if excel_file:
                report_files['excel'] = excel_file
        
        self.logger.info(f"Generated {len(report_files)} report files")
        return report_files
    
    def generate_html_report(self, data_dict: Dict[str, pd.DataFrame], 
                           comparison_df: pd.DataFrame,
                           stats_dict: Dict[str, Dict[str, float]],
                           timestamp: datetime) -> Optional[str]:
        """Generate interactive HTML report."""
        try:
            # Create charts
            charts = self._create_interactive_charts(data_dict, comparison_df)
            
            # Get stock symbols
            symbols = list(data_dict.keys())
            
            # Prepare template data
            template_data = {
                'title': 'Stock Analysis Report',
                'subtitle': f'Analysis of {", ".join(symbols)}',
                'generation_time': timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'analysis_type': 'Multi-Stock Analysis',
                'symbols': symbols,
                'charts': charts,
                'statistics': stats_dict,
                'data_summary': self._create_data_summary(data_dict)
            }
            
            # Render HTML template
            html_content = self._render_html_template(template_data)
            
            # Save HTML file
            filename = f"stock_analysis_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.html"
            filepath = os.path.join(Config.REPORTS_DIR, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"Generated HTML report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate HTML report: {e}")
            return None
    
    def generate_pdf_report(self, data_dict: Dict[str, pd.DataFrame], 
                          comparison_df: pd.DataFrame,
                          stats_dict: Dict[str, Dict[str, float]],
                          timestamp: datetime) -> Optional[str]:
        """Generate PDF report."""
        if not WEASYPRINT_AVAILABLE:
            self.logger.warning("WeasyPrint not available - skipping PDF generation")
            return None
            
        try:
            # Create static charts for PDF
            charts = self._create_static_charts(data_dict, comparison_df)
            
            # Get stock symbols
            symbols = list(data_dict.keys())
            
            # Prepare template data
            template_data = {
                'title': 'Stock Analysis Report',
                'subtitle': f'Analysis of {", ".join(symbols)}',
                'generation_time': timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'analysis_type': 'Multi-Stock Analysis',
                'symbols': symbols,
                'charts': charts,
                'statistics': stats_dict,
                'data_summary': self._create_data_summary(data_dict),
                'key_insights': self._generate_key_insights(stats_dict, comparison_df)
            }
            
            # Render HTML template for PDF
            html_content = self._render_pdf_template(template_data)
            
            # Convert to PDF
            filename = f"stock_analysis_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(Config.REPORTS_DIR, filename)
            
            weasyprint.HTML(string=html_content).write_pdf(filepath)
            
            self.logger.info(f"Generated PDF report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate PDF report: {e}")
            return None
    
    def generate_excel_report(self, data_dict: Dict[str, pd.DataFrame], 
                            comparison_df: pd.DataFrame,
                            stats_dict: Dict[str, Dict[str, float]],
                            timestamp: datetime) -> Optional[str]:
        """Generate Excel report with multiple sheets."""
        try:
            filename = f"stock_analysis_report_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(Config.REPORTS_DIR, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Summary sheet
                self._create_summary_sheet(writer, stats_dict, timestamp)
                
                # Individual stock sheets
                for symbol, df in data_dict.items():
                    if not df.empty:
                        self._create_stock_sheet(writer, df, symbol)
                
                # Comparison sheet
                if not comparison_df.empty:
                    self._create_comparison_sheet(writer, comparison_df)
                
                # Technical analysis sheet
                self._create_technical_analysis_sheet(writer, data_dict)
            
            # Apply additional formatting
            self._format_excel_file(filepath)
            
            self.logger.info(f"Generated Excel report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            return None
    
    def _create_interactive_charts(self, data_dict: Dict[str, pd.DataFrame], 
                                 comparison_df: pd.DataFrame) -> Dict[str, str]:
        """Create interactive Plotly charts for HTML report."""
        charts = {}
        
        # Price comparison chart
        fig_price = make_subplots(
            rows=2, cols=1,
            subplot_titles=('Price Comparison', 'Volume Comparison'),
            vertical_spacing=0.1,
            row_heights=[0.7, 0.3]
        )
        
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
        
        for i, (symbol, df) in enumerate(data_dict.items()):
            if not df.empty and 'close' in df.columns:
                fig_price.add_trace(
                    go.Scatter(
                        x=df.index,
                        y=df['close'],
                        name=f'{symbol} Price',
                        line=dict(color=colors[i % len(colors)], width=2),
                        hovertemplate=f'{symbol}<br>Date: %{{x}}<br>Price: $%{{y:.2f}}<extra></extra>'
                    ),
                    row=1, col=1
                )
                
                if 'volume' in df.columns:
                    fig_price.add_trace(
                        go.Bar(
                            x=df.index,
                            y=df['volume'],
                            name=f'{symbol} Volume',
                            marker_color=colors[i % len(colors)],
                            opacity=0.6,
                            hovertemplate=f'{symbol}<br>Date: %{{x}}<br>Volume: %{{y:,.0f}}<extra></extra>'
                        ),
                        row=2, col=1
                    )
        
        fig_price.update_layout(
            title="Stock Price and Volume Analysis",
            height=600,
            template=Config.CHART_STYLE,
            hovermode='x unified'
        )
        fig_price.update_xaxes(title_text="Date", row=2, col=1)
        fig_price.update_yaxes(title_text="Price ($)", row=1, col=1)
        fig_price.update_yaxes(title_text="Volume", row=2, col=1)
        
        charts['price_comparison'] = fig_price.to_html(full_html=False, include_plotlyjs='cdn')
        
        # Technical indicators chart
        if not data_dict:
            return charts
        
        fig_technical = make_subplots(
            rows=3, cols=1,
            subplot_titles=('Price with Moving Averages', 'RSI', 'MACD'),
            vertical_spacing=0.08,
            row_heights=[0.5, 0.25, 0.25]
        )
        
        # Add price and moving averages
        for i, (symbol, df) in enumerate(data_dict.items()):
            if not df.empty and 'close' in df.columns:
                fig_technical.add_trace(
                    go.Scatter(
                        x=df.index,
                        y=df['close'],
                        name=f'{symbol} Price',
                        line=dict(color=colors[i % len(colors)], width=2)
                    ),
                    row=1, col=1
                )
                
                # Add moving averages
                if 'sma_20' in df.columns:
                    fig_technical.add_trace(
                        go.Scatter(
                            x=df.index,
                            y=df['sma_20'],
                            name=f'{symbol} SMA 20',
                            line=dict(color=colors[i % len(colors)], width=1, dash='dash'),
                            opacity=0.7
                        ),
                        row=1, col=1
                    )
                
                # Add RSI
                if 'rsi' in df.columns:
                    fig_technical.add_trace(
                        go.Scatter(
                            x=df.index,
                            y=df['rsi'],
                            name=f'{symbol} RSI',
                            line=dict(color=colors[i % len(colors)], width=2)
                        ),
                        row=2, col=1
                    )
                
                # Add MACD
                if 'macd' in df.columns:
                    fig_technical.add_trace(
                        go.Scatter(
                            x=df.index,
                            y=df['macd'],
                            name=f'{symbol} MACD',
                            line=dict(color=colors[i % len(colors)], width=2)
                        ),
                        row=3, col=1
                    )
        
        # Add RSI reference lines
        fig_technical.add_hline(y=70, line_dash="dash", line_color="red", opacity=0.5, row=2, col=1)
        fig_technical.add_hline(y=30, line_dash="dash", line_color="green", opacity=0.5, row=2, col=1)
        fig_technical.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5, row=3, col=1)
        
        fig_technical.update_layout(
            title="Technical Analysis",
            height=800,
            template=Config.CHART_STYLE,
            hovermode='x unified'
        )
        
        charts['technical_analysis'] = fig_technical.to_html(full_html=False, include_plotlyjs='cdn')
        
        # Performance comparison chart
        if not comparison_df.empty:
            symbols = list(data_dict.keys())
            if len(symbols) >= 2:
                normalized_cols = [f'{sym}_close_normalized' for sym in symbols if f'{sym}_close_normalized' in comparison_df.columns]
                
                if normalized_cols:
                    fig_performance = go.Figure()
                    
                    for i, col in enumerate(normalized_cols):
                        symbol = col.replace('_close_normalized', '')
                        fig_performance.add_trace(
                            go.Scatter(
                                x=comparison_df.index,
                                y=comparison_df[col],
                                name=f'{symbol} Performance',
                                line=dict(color=colors[i % len(colors)], width=3),
                                hovertemplate=f'{symbol}<br>Date: %{{x}}<br>Performance: %{{y:.1f}}%<extra></extra>'
                            )
                        )
                    
                    fig_performance.update_layout(
                        title="Relative Performance Comparison (Normalized to 100)",
                        xaxis_title="Date",
                        yaxis_title="Performance Index",
                        height=500,
                        template=Config.CHART_STYLE,
                        hovermode='x unified'
                    )
                    
                    charts['performance_comparison'] = fig_performance.to_html(full_html=False, include_plotlyjs='cdn')
        
        return charts
    
    def _create_static_charts(self, data_dict: Dict[str, pd.DataFrame], 
                            comparison_df: pd.DataFrame) -> Dict[str, str]:
        """Create static charts for PDF report."""
        charts = {}
        
        # Similar logic to interactive charts but save as base64 encoded images
        # This is a simplified version - you can expand based on needs
        
        try:
            # Price comparison chart
            fig_price = go.Figure()
            
            colors = ['#1f77b4', '#ff7f0e']
            for i, (symbol, df) in enumerate(data_dict.items()):
                if not df.empty and 'close' in df.columns:
                    fig_price.add_trace(
                        go.Scatter(
                            x=df.index,
                            y=df['close'],
                            name=f'{symbol}',
                            line=dict(color=colors[i % len(colors)], width=2)
                        )
                    )
            
            fig_price.update_layout(
                title="Stock Price Comparison",
                xaxis_title="Date",
                yaxis_title="Price ($)",
                height=400,
                template=Config.CHART_STYLE
            )
            
            # Convert to base64 for embedding in PDF
            try:
                img_bytes = fig_price.to_image(format="png", width=800, height=400)
                img_base64 = base64.b64encode(img_bytes).decode()
                charts['price_comparison'] = f"data:image/png;base64,{img_base64}"
            except Exception as img_error:
                self.logger.warning(f"Failed to generate chart image: {img_error}")
                charts['price_comparison'] = "Chart generation not available"
            
        except Exception as e:
            self.logger.warning(f"Failed to create static chart: {e}")
        
        return charts
    
    def _render_html_template(self, template_data: Dict[str, Any]) -> str:
        """Render HTML template."""
        # Create a basic HTML template if none exists
        html_template = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{{ title }}</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
                .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }
                .title { font-size: 2.5rem; margin: 0; }
                .subtitle { font-size: 1.2rem; margin: 10px 0 0 0; opacity: 0.9; }
                .metadata { font-size: 0.9rem; margin-top: 15px; opacity: 0.8; }
                .content { background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 20px; }
                .section-title { font-size: 1.8rem; color: #333; margin-bottom: 20px; border-bottom: 2px solid #667eea; padding-bottom: 10px; }
                .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 20px 0; }
                .stats-card { background: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 4px solid #667eea; }
                .stats-card h3 { color: #333; margin-top: 0; }
                .metric { display: flex; justify-content: space-between; margin: 10px 0; }
                .metric-name { font-weight: 500; }
                .metric-value { font-weight: bold; color: #667eea; }
                .chart-container { margin: 30px 0; }
                .positive { color: #28a745; }
                .negative { color: #dc3545; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1 class="title">{{ title }}</h1>
                <p class="subtitle">{{ subtitle }}</p>
                <div class="metadata">
                    Generated on {{ generation_time }} | Analysis Type: {{ analysis_type }}
                </div>
            </div>
            
            <div class="content">
                <h2 class="section-title">Summary Statistics</h2>
                <div class="stats-grid">
                    {% for symbol, stats in statistics.items() %}
                    <div class="stats-card">
                        <h3>{{ symbol }} Statistics</h3>
                        {% for metric, value in stats.items() %}
                        <div class="metric">
                            <span class="metric-name">{{ metric.replace('_', ' ').title() }}:</span>
                            <span class="metric-value {% if value > 0 and 'change' in metric %}positive{% elif value < 0 and 'change' in metric %}negative{% endif %}">
                                {% if 'price' in metric %}${{ "%.2f"|format(value) }}{% elif 'pct' in metric %}{{ "%.2f"|format(value) }}%{% else %}{{ "%.2f"|format(value) }}{% endif %}
                            </span>
                        </div>
                        {% endfor %}
                    </div>
                    {% endfor %}
                </div>
            </div>
            
            <div class="content">
                <h2 class="section-title">Charts and Analysis</h2>
                {% for chart_name, chart_html in charts.items() %}
                <div class="chart-container">
                    <h3>{{ chart_name.replace('_', ' ').title() }}</h3>
                    {{ chart_html|safe }}
                </div>
                {% endfor %}
            </div>
        </body>
        </html>
        """
        
        template = Template(html_template)
        return template.render(**template_data)
    
    def _render_pdf_template(self, template_data: Dict[str, Any]) -> str:
        """Render HTML template optimized for PDF conversion."""
        # Simplified template for PDF
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; font-size: 12px; }
                .header { text-align: center; margin-bottom: 30px; border-bottom: 2px solid #333; padding-bottom: 20px; }
                .title { font-size: 24px; margin: 0; color: #333; }
                .subtitle { font-size: 16px; margin: 10px 0; color: #666; }
                .section { margin: 30px 0; }
                .section-title { font-size: 18px; color: #333; margin-bottom: 15px; border-bottom: 1px solid #ccc; padding-bottom: 5px; }
                .stats-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                .stats-table th, .stats-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                .stats-table th { background-color: #f2f2f2; }
                .insight { background: #f8f9fa; padding: 15px; border-left: 4px solid #007bff; margin: 15px 0; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1 class="title">{{ title }}</h1>
                <p class="subtitle">{{ subtitle }}</p>
                <p>Generated on {{ generation_time }}</p>
            </div>
            
            <div class="section">
                <h2 class="section-title">Executive Summary</h2>
                {% for insight in key_insights %}
                <div class="insight">{{ insight }}</div>
                {% endfor %}
            </div>
            
            <div class="section">
                <h2 class="section-title">Key Metrics</h2>
                <table class="stats-table">
                    <thead>
                        <tr><th>Symbol</th><th>Current Price</th><th>1D Change</th><th>1D Change %</th><th>Volatility</th></tr>
                    </thead>
                    <tbody>
                        {% for symbol, stats in statistics.items() %}
                        <tr>
                            <td>{{ symbol }}</td>
                            <td>${{ "%.2f"|format(stats.get('current_price', 0)) }}</td>
                            <td>${{ "%.2f"|format(stats.get('price_change_1d', 0)) }}</td>
                            <td>{{ "%.2f"|format(stats.get('price_change_pct_1d', 0)) }}%</td>
                            <td>{{ "%.2f"|format(stats.get('volatility', 0)) }}%</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </body>
        </html>
        """
        
        template = Template(html_template)
        return template.render(**template_data)
    
    def _create_data_summary(self, data_dict: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Create data summary for reports."""
        summary = {
            'total_symbols': len(data_dict),
            'data_points': sum(len(df) for df in data_dict.values()),
            'date_range': None
        }
        
        if data_dict:
            all_dates = []
            for df in data_dict.values():
                if not df.empty:
                    all_dates.extend(df.index.tolist())
            
            if all_dates:
                summary['date_range'] = {
                    'start': min(all_dates).strftime('%Y-%m-%d'),
                    'end': max(all_dates).strftime('%Y-%m-%d')
                }
        
        return summary
    
    def _generate_key_insights(self, stats_dict: Dict[str, Dict[str, float]], 
                             comparison_df: pd.DataFrame) -> List[str]:
        """Generate key insights for the report."""
        insights = []
        
        if len(stats_dict) >= 2:
            symbols = list(stats_dict.keys())
            
            # Price performance insight
            price_changes = {symbol: stats.get('price_change_pct_1d', 0) 
                           for symbol, stats in stats_dict.items()}
            best_performer = max(price_changes, key=price_changes.get)
            worst_performer = min(price_changes, key=price_changes.get)
            
            insights.append(f"{best_performer} outperformed {worst_performer} with a {price_changes[best_performer]:.2f}% daily change vs {price_changes[worst_performer]:.2f}%")
            
            # Volatility insight
            volatilities = {symbol: stats.get('volatility', 0) 
                          for symbol, stats in stats_dict.items()}
            most_volatile = max(volatilities, key=volatilities.get)
            least_volatile = min(volatilities, key=volatilities.get)
            
            insights.append(f"{most_volatile} shows higher volatility ({volatilities[most_volatile]:.2f}%) compared to {least_volatile} ({volatilities[least_volatile]:.2f}%)")
        
        return insights
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, stats_dict: Dict[str, Dict[str, float]], 
                            timestamp: datetime) -> None:
        """Create summary sheet in Excel."""
        summary_data = []
        
        for symbol, stats in stats_dict.items():
            summary_data.append({
                'Symbol': symbol,
                'Current Price': stats.get('current_price', 0),
                '1D Change': stats.get('price_change_1d', 0),
                '1D Change %': stats.get('price_change_pct_1d', 0),
                '7D Change %': stats.get('price_change_7d', 0),
                '30D Change %': stats.get('price_change_30d', 0),
                'Volatility %': stats.get('volatility', 0),
                'RSI': stats.get('current_rsi', 0)
            })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_stock_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, symbol: str) -> None:
        """Create individual stock sheet in Excel."""
        # Select key columns for Excel
        columns_to_include = ['open', 'high', 'low', 'close', 'volume']
        if 'daily_return' in df.columns:
            columns_to_include.append('daily_return')
        if 'sma_20' in df.columns:
            columns_to_include.append('sma_20')
        if 'rsi' in df.columns:
            columns_to_include.append('rsi')
        
        df_export = df[columns_to_include].copy()
        
        # Remove timezone from datetime index for Excel compatibility
        if isinstance(df_export.index, pd.DatetimeIndex) and df_export.index.tz is not None:
            df_export.index = df_export.index.tz_localize(None)
        
        df_export.to_excel(writer, sheet_name=f'{symbol}_Data')
    
    def _create_comparison_sheet(self, writer: pd.ExcelWriter, comparison_df: pd.DataFrame) -> None:
        """Create comparison sheet in Excel."""
        # Remove timezone from datetime index for Excel compatibility
        comparison_excel = comparison_df.copy()
        if isinstance(comparison_excel.index, pd.DatetimeIndex) and comparison_excel.index.tz is not None:
            comparison_excel.index = comparison_excel.index.tz_localize(None)
        
        comparison_excel.to_excel(writer, sheet_name='Comparison')
    
    def _create_technical_analysis_sheet(self, writer: pd.ExcelWriter, data_dict: Dict[str, pd.DataFrame]) -> None:
        """Create technical analysis sheet in Excel."""
        tech_data = []
        
        for symbol, df in data_dict.items():
            if not df.empty and 'close' in df.columns:
                latest_data = df.iloc[-1]
                tech_data.append({
                    'Symbol': symbol,
                    'Current Price': latest_data.get('close', 0),
                    'SMA 20': latest_data.get('sma_20', 0),
                    'SMA 50': latest_data.get('sma_50', 0),
                    'RSI': latest_data.get('rsi', 0),
                    'MACD': latest_data.get('macd', 0),
                    'BB Upper': latest_data.get('bb_upper', 0),
                    'BB Lower': latest_data.get('bb_lower', 0),
                })
        
        df_tech = pd.DataFrame(tech_data)
        df_tech.to_excel(writer, sheet_name='Technical_Analysis', index=False)
    
    def _format_excel_file(self, filepath: str) -> None:
        """Apply additional formatting to Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import NamedStyle
            
            wb = load_workbook(filepath)
            
            # Define styles
            header_style = NamedStyle(name="header")
            header_style.font = Font(bold=True, color="FFFFFF")
            header_style.fill = PatternFill("solid", fgColor="366092")
            header_style.alignment = Alignment(horizontal="center")
            
            # Apply to all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Apply header style to first row
                for cell in ws[1]:
                    cell.style = header_style
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            
        except Exception as e:
            self.logger.warning(f"Failed to apply Excel formatting: {e}") 